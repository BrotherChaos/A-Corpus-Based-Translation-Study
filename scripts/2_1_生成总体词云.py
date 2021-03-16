    
from wordcloud import WordCloud
import openpyxl

def get_text(sheet_name):

    wb = openpyxl.load_workbook("试验数据\meta_data_0309.xlsx")
    sheet = wb.get_sheet_by_name(sheet_name)
    idiom_list = []
    max_row = sheet.max_row
    for i in range(max_row-1):
        idiom_list.append(sheet["A{}".format(i+2)].value)
    text = " ".join(idiom_list)
    return text,sheet_name



def generate_wordcloud(text):

    wc = WordCloud(
    
    # 设置字体，不指定就会出现乱码
    # font_path="C:\\Windows\\Fonts\\STFANGSO.ttf",
    font_path="C:\\Windows\\Fonts\\msyh.ttc",
    # 设置背景色
    background_color='white',
    # 设置背景宽
    width=1000,
    # 设置背景高
    height=500,
    # 最大字体
        max_font_size=70,
    # 最小字体
    min_font_size=15,
    mode='RGBA',
    collocations=False,
    max_words = 100
    )
    # 产生词云
    wc.generate(text)
    # 保存图片
    wc.to_file("试验数据\总词云图_max100.png") # 按照设置的像素宽高度保存绘制好的词云图，比下面程序显示更清晰
   
f0 = openpyxl.load_workbook("试验数据\meta_data_0309.xlsx")
sheets_list= f0.get_sheet_names()
sum_txt = ""
for i in sheets_list:
    text,file_name = get_text(i)
    sum_txt += text
print(sum_txt)
generate_wordcloud(sum_txt)