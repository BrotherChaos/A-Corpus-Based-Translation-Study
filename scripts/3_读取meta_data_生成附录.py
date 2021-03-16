import openpyxl
wb = openpyxl.load_workbook("试验数据\meta_data_0309.xlsx")
app = openpyxl.load_workbook("试验数据\附录_0313.xlsx")
sheet_name_list = wb.get_sheet_names()
for sheet_name in sheet_name_list:
    idiom_sheet = wb.get_sheet_by_name(sheet_name)
    newsheet = app.create_sheet(sheet_name)
    idiom_list = []
    count_dic = {}
    for i in range(idiom_sheet.max_row):
        idiom_list.append(idiom_sheet["A{}".format(i+1)].value)
    a = set(idiom_list)
    for i in a:
        count_dic[i] = idiom_list.count(i)

    idiom = []
    num = []
    for key,value in count_dic.items():
        idiom.append(key)
        num.append(value)
    for i in range(len(idiom)):
        print(idiom[i])
        print(num[i])
    
        newsheet["A{}".format(i+1)] = idiom[i]
        newsheet["B{}".format(i+1)] = num[i]
app.save("试验数据\附录_0313.xlsx")  
        




